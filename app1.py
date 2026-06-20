import os
import sys
import shutil
import subprocess
import hashlib

from datetime import datetime

import tkinter as tk

from tkinter import (
    ttk,
    messagebox,
    simpledialog
)

# ==========================================
# НАСТРОЙКИ (ЗАПОЛНЯЮТСЯ КОНФИГУРАТОРОМ)
# ==========================================
EXCEL_FILE = {{ EXCEL_FILE }}
SECRET_DB_FOLDER = {{ SECRET_DB_FOLDER }}
TEMP_DIR = {{ TEMP_DIR }}
ADMIN_PASSWORD = "{{ ADMIN_PASSWORD }}"
# ==========================================

# ==========================================
# LAZY IMPORTS
# ==========================================

_sqlite = None

_load_workbook = None

_get_column_letter = None


def get_sqlite():
    global _sqlite
    if _sqlite is None:
        import sqlite3
        _sqlite = sqlite3

        print(
            "sqlite загружен"
        )
    return _sqlite


def get_load_workbook():

    global _load_workbook

    if _load_workbook is None:

        from openpyxl import (
            load_workbook
        )

        _load_workbook = (
            load_workbook
        )

        print(
            "openpyxl загружен"
        )

    return _load_workbook


def get_column():

    global _get_column_letter

    if _get_column_letter is None:

        from openpyxl.utils import (
            get_column_letter
        )

        _get_column_letter = (
            get_column_letter
        )

    return _get_column_letter

def preload_modules():

    try:

        get_sqlite()

        get_load_workbook()

        get_column()

        print(
            "Модули догружены"
        )

    except Exception as e:

        print(
            f"Ошибка preload: {e}"
        )

# ---------- КЭШ БУКВ ----------

_COL_LETTER_CACHE = {}

def col_letter(col_num):
    if col_num not in _COL_LETTER_CACHE:
        _COL_LETTER_CACHE[
            col_num
        ] = (
            get_column()(
                col_num
            )
        )
    return (
        _COL_LETTER_CACHE[
            col_num
        ]
    )

# ---------- ПУТИ ----------

excel_dir = os.path.dirname(EXCEL_FILE)
excel_name = os.path.splitext(os.path.basename(EXCEL_FILE))[0]
LOCK_FILE = os.path.join(excel_dir, f".{excel_name}.lock")

# Основная БД – сначала пытаемся создать рядом с Excel

DB_PATH = os.path.join(excel_dir, "excel_audit.db")
try:
    test_file = os.path.join(excel_dir, ".write_test")
    with open(test_file, 'w') as f:
        f.write('test')
    os.remove(test_file)
except:
    DB_PATH = os.path.join(os.environ.get('APPDATA', ''), "excel_audit.db")
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

# Резервная БД (секретная) – без сообщений
def get_secret_db_path():
    global SECRET_DB_FOLDER
    if not os.path.exists(SECRET_DB_FOLDER):
        try:
            os.makedirs(SECRET_DB_FOLDER, exist_ok=True)
        except:
            fallback = os.path.join(os.environ.get('APPDATA', ''), 'ExcelAudit')
            os.makedirs(fallback, exist_ok=True)
            SECRET_DB_FOLDER = fallback
    return os.path.join(SECRET_DB_FOLDER, "excel_audit.db")

BACKUP_DB = get_secret_db_path()

# ---------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------

def get_user():
    return os.environ.get("USERNAME", "UNKNOWN")

def get_lock_info():
    if not os.path.exists(LOCK_FILE):
        return None
    try:
        with open(LOCK_FILE, 'r') as f:
            owner = f.readline().strip()
            start_str = f.readline().strip()
            start_time = datetime.fromisoformat(start_str)
            return owner, start_time
    except:
        return None

def is_lock_stale():
    info = get_lock_info()
    if info is None:
        return False
    _, start_time = info
    minutes_passed = (datetime.now() - start_time).total_seconds() / 60
    return minutes_passed > 30

def force_remove_stale_lock():
    if is_lock_stale():
        try:
            os.remove(LOCK_FILE)
            return True
        except:
            pass
    return False

def file_hash():
    h = hashlib.blake2b()
    with open(EXCEL_FILE, "rb") as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()

# ---------- БАЗА ДАННЫХ ----------

def init_db():
    global DB_PATH
    db_dir = os.path.dirname(DB_PATH)
    if not os.path.exists(db_dir):
        try:
            os.makedirs(db_dir, exist_ok=True)
        except:
            DB_PATH = os.path.join(os.environ.get('APPDATA', ''), 'excel_audit.db')
            db_dir = os.path.dirname(DB_PATH)
            os.makedirs(db_dir, exist_ok=True)
    else:
        test_file = os.path.join(db_dir, '.write_test')
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
        except:
            DB_PATH = os.path.join(os.environ.get('APPDATA', ''), 'excel_audit.db')
            db_dir = os.path.dirname(DB_PATH)
            os.makedirs(db_dir, exist_ok=True)

    conn = get_sqlite().connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS main_state (
            sheet TEXT,
            row_num INTEGER,
            col_num INTEGER,
            value TEXT,
            col_letter TEXT,
            col_title TEXT,
            row_title TEXT,
            PRIMARY KEY (sheet, row_num, col_num)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_time TEXT,
            user_name TEXT,
            event_type TEXT,
            sheet TEXT,
            col_letter TEXT,
            col_title TEXT,
            row_num INTEGER,
            row_title TEXT,
            old_value TEXT,
            new_value TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_time ON audit(event_time)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_user ON audit(user_name)")
    conn.commit()
    conn.close()

def backup_db():
    if os.path.exists(DB_PATH):
        shutil.copy2(DB_PATH, BACKUP_DB)

def add_audit_records(records):
    if not records:
        return
    conn = get_sqlite().connect(DB_PATH)
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA journal_mode=WAL")
    now = datetime.now().isoformat()
    user = get_user()
    rows = [(now, user) + rec for rec in records]
    conn.executemany("""
        INSERT INTO audit(
            event_time, user_name, event_type,
            sheet, col_letter, col_title,
            row_num, row_title, old_value, new_value
        ) VALUES(?,?,?,?,?,?,?,?,?,?)
    """, rows)
    conn.commit()
    conn.close()
    if len(records) > 100:
        backup_db()

# ---------- СКАНИРОВАНИЕ И АУДИТ ----------

def full_scan(parent):
    pwd = simpledialog.askstring("Пароль", "Введите пароль администратора:", parent=parent, show='*')
    if pwd != ADMIN_PASSWORD:
        messagebox.showerror("Ошибка", "Неверный пароль!")
        return

    wb = get_load_workbook()(EXCEL_FILE, read_only=True, data_only=False)
    sheets = wb.sheetnames
    total = len(sheets)

    progress_win = tk.Toplevel(parent)
    progress_win.title("Сканирование Excel")
    progress_win.geometry("400x120")
    progress_win.transient(parent)
    progress_win.grab_set()
    label = tk.Label(progress_win, text="Подготовка...")
    label.pack(pady=10)
    progress = ttk.Progressbar(progress_win, maximum=total, mode='determinate')
    progress.pack(pady=10, padx=20, fill='x')
    parent.update()

    main_records = []
    audit_records = []

    for idx, sheet_name in enumerate(sheets):
        label.config(text=f"Лист {idx+1}/{total}: {sheet_name}")
        progress['value'] = idx
        parent.update()
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            continue
        for r_num, row in enumerate(rows_iter, start=2):
            if not row:
                continue
            row_title = str(row[0]) if row[0] is not None else ""
            for c_num, value in enumerate(row, start=1):
                if c_num == 1 or value is None:
                    continue
                col_letter_str = col_letter(c_num)
                col_title = str(headers[c_num-1]) if c_num-1 < len(headers) and headers[c_num-1] is not None else ""
                str_value = str(value)
                main_records.append((sheet_name, r_num, c_num, str_value, col_letter_str, col_title, row_title))
                audit_records.append(("INITIAL_RECORD", sheet_name, col_letter_str, col_title,
                                      r_num, row_title, "", str_value))
    wb.close()
    progress_win.destroy()

    conn = get_sqlite().connect(DB_PATH)
    conn.execute("DROP TABLE IF EXISTS main_state")
    conn.close()
    init_db()

    conn = get_sqlite().connect(DB_PATH)
    cur = conn.cursor()
    cur.executemany("INSERT INTO main_state VALUES(?,?,?,?,?,?,?)", main_records)
    conn.commit()
    conn.close()

    if audit_records:
        add_audit_records(audit_records)

    messagebox.showinfo("Готово", f"База создана ({len(main_records)} ячеек)")

def open_excel_and_wait(file_path):

    file_path = os.path.abspath(file_path)

    try:

        subprocess.run(
            f'cmd /c start /wait excel.exe "{file_path}"',
            shell=True,
            check=True
        )

        return

    except:

        excel_paths = [

            r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",

            r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE"

        ]

        for excel in excel_paths:

            try:

                subprocess.run(
                    [excel, file_path],
                    check=True
                )

                return

            except:

                pass

    raise RuntimeError(
        "Microsoft Excel не найден"
    )

def lock():
    if os.path.exists(LOCK_FILE):
        if is_lock_stale():
            try:
                os.remove(
                    LOCK_FILE
                )
            except:
                return False
        else:
            return False
    with open(
        LOCK_FILE,
        "w",
        encoding="utf-8"
    ) as f:
        f.write(
            get_user()
        )
        f.write("\n")
        f.write(
            datetime.now().isoformat()
        )
        f.write("\n")
        f.write(
            "EDIT_OR_UPDATE"
        )
    return True

def unlock():
    try:
        if os.path.exists(
            LOCK_FILE
        ):
            os.remove(
                LOCK_FILE
            )
    except:
        pass

def make_copy():

    temp = os.path.abspath(TEMP_DIR)

    os.makedirs(
        temp,
        exist_ok=True
    )

    dst = os.path.join(
        temp,
        os.path.basename(EXCEL_FILE)
    )

    shutil.copy2(
        EXCEL_FILE,
        dst
    )

    return dst

def open_copy():

    copy_path = make_copy()

    folder = os.path.dirname(copy_path)

    if os.path.exists(folder):

        subprocess.Popen([
            "explorer",
            os.path.abspath(folder)
        ])

    else:

        messagebox.showerror(
            "Ошибка",
            f"Не найдена папка:\n{folder}"
        )

def background_audit():
    if not lock():
        return

    try:
        print("Открытие Excel...")
        # Предзагрузка модулей (не блокирует)
        import threading
        thread = threading.Thread(target=preload_modules, daemon=True)
        thread.start()

        open_excel_and_wait(EXCEL_FILE)
        thread.join()
        print("Excel закрыт")

        print("Начинается аудит...")
        wb = get_load_workbook()(EXCEL_FILE, read_only=True, data_only=False)
        sheets = wb.sheetnames
        current_state = []

        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                continue
            for r_num, row in enumerate(rows_iter, start=2):
                if not row:
                    continue
                row_title = str(row[0]) if row[0] is not None else ""
                for c_num, value in enumerate(row, start=1):
                    if c_num == 1 or value is None:
                        continue
                    col_letter_str = col_letter(c_num)
                    col_title = str(headers[c_num-1]) if c_num-1 < len(headers) and headers[c_num-1] is not None else ""
                    current_state.append((sheet_name, r_num, c_num, str(value), col_letter_str, col_title, row_title))
        wb.close()
        print("Сканирование завершено")

        # Получаем старый эталон из БД
        conn = get_sqlite().connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT sheet, row_num, col_num, value, col_letter, col_title, row_title FROM main_state")
        old_rows = cur.fetchall()
        old_dict = {(s, r, c): (v, cl, ct, rt) for (s, r, c, v, cl, ct, rt) in old_rows}

        changes = []
        new_state_dict = {}
        for (sheet, row, col, new_val, col_letter_str, col_title, row_title) in current_state:
            key = (sheet, row, col)
            old_info = old_dict.get(key)
            if old_info is None:
                changes.append(("CELL_CHANGE", sheet, col_letter_str, col_title,
                                row, row_title, "", new_val))
            else:
                old_val, _, _, _ = old_info
                if old_val != new_val:
                    changes.append(("CELL_CHANGE", sheet, col_letter_str, col_title,
                                    row, row_title, old_val, new_val))
            new_state_dict[key] = (sheet, row, col, new_val, col_letter_str, col_title, row_title)

        # Удалённые ячейки
        for key, (old_val, cl, ct, rt) in old_dict.items():
            if key not in new_state_dict:
                sheet, row, col = key
                changes.append(("CELL_CHANGE", sheet, cl, ct,
                                row, rt, old_val, ""))

        if changes:
            add_audit_records(changes)
            print(f"Записано {len(changes)} изменений")
        else:
            print("Изменений не найдено")

        # Обновляем эталон
        cur.execute("BEGIN TRANSACTION")
        data = list(new_state_dict.values())
        cur.executemany("""
            INSERT OR REPLACE INTO main_state
            (sheet, row_num, col_num, value, col_letter, col_title, row_title)
            VALUES(?,?,?,?,?,?,?)
        """, data)
        to_delete = list(old_dict.keys() - new_state_dict.keys())
        if to_delete:
            cur.executemany("DELETE FROM main_state WHERE sheet=? AND row_num=? AND col_num=?", to_delete)
        conn.commit()
        conn.close()
        print("База обновлена")

    finally:
        print("Снятие блокировки")
        unlock()

def check_has_main():

    try:

        conn = get_sqlite().connect(DB_PATH)

        cur = conn.cursor()

        cur.execute(
            """
            SELECT count(*)
            FROM sqlite_master
            WHERE type='table'
            AND name='main_state'
            """
        )

        if cur.fetchone()[0] == 0:

            conn.close()

            return False

        cur.execute(
            "SELECT count(*) FROM main_state"
        )

        count = cur.fetchone()[0]

        conn.close()

        return count > 0

    except:

        return False

# ---------- GUI ----------
class App:
    def __init__(self, root):
        self.root = root
        root.title("Аудит Excel")
        root.geometry("400x300")
        root.resizable(False, False)

        self.btn_edit = tk.Button(root, text="✏️ ОТКРЫТЬ ОРИГИНАЛ", height=3,
                                  command=self.on_edit_original)
        self.btn_edit.pack(fill="x", padx=20, pady=10)

        self.btn_copy = tk.Button(root, text="📋 ПОЛУЧИТЬ КОПИЮ", height=3,
                                  command=self.on_copy)
        self.btn_copy.pack(fill="x", padx=20, pady=10)

        self.btn_admin = tk.Button(root, text="Создать базу (админ)", command=lambda: full_scan(root))
        self.btn_admin.pack(pady=(20, 5))

        self.status_label = tk.Label(root, text="", fg="gray")
        self.status_label.pack()
        self.update_status()

    def update_status(self):
        try:
            state = check_has_main()
        except:
            state = False

        if state:
            self.status_label.config(text="✅ База создана", fg="green")
        else:
            self.status_label.config(text="⚠️ База не создана. Нажмите 'Создать базу'", fg="red")
        self.root.after(5000, self.update_status)

    def on_edit_original(self):

        if os.path.exists(LOCK_FILE):

            info = get_lock_info()

            if info:

                owner, _ = info

                messagebox.showwarning(
                    "Файл занят",
                    (
                        "Сейчас выполняется обновление базы.\n\n"
                        f"Пользователь:\n{owner}"
                    )
                )

            else:

                messagebox.showwarning(
                    "Файл занят",
                    "Подождите завершения обновления"
                )

            return

        self.root.withdraw()

        try:

            background_audit()

            self.root.destroy()

        except Exception as e:

            self.root.deiconify()

            messagebox.showerror(
                "Ошибка",
                str(e)
            )

    def on_copy(self):
        open_copy()
        self.root.destroy()

# ---------- ЗАПУСК ----------
if __name__ == "__main__":

    try:
        init_db()

    except Exception as e:

        print(
            f"Ошибка инициализации БД: {e}"
        )

    if (
        len(sys.argv) > 1
        and
        sys.argv[1] == "--background-original"
    ):

        background_audit()

        sys.exit(0)

    root = tk.Tk()

    app = App(root)

    root.mainloop()
